from openpyxl import load_workbook
import os.path

path = os.path.abspath(os.path.dirname(__file__))
datapath = os.path.join(path, '../data/tyre-dealers-name-and-address.xlsx')
outpath = os.path.join(path, '../output/tyre-care.partial.html')

wb = load_workbook(filename = datapath)
ws = wb.active

with open(outpath, 'w') as f:
    for i in range(2, len(ws['A'])):
        try:
            dealer_name = (ws.cell(row=i+1, column=1).value).title()
            dealer_address = (ws.cell(row=i+1, column=2).value).title()
            f.write(f"""
<div class="col-12 col-md-4 my-3">
    <div class="card">
        <div class="card-body">
            <h5 class="card-title">{dealer_name}</h5>
            <p class="card-text"><i class="icon-location-pin"></i>&ensp;{dealer_address}</p>
        </div>
        <div class="card-footer text-center">
            <a href="/enquiry.html?tyre_dealer={dealer_name}" class="btn btn-outline-primary">Contact Dealer</a>
        </div>
    </div>
</div>\n""")
        except AttributeError:
            pass

